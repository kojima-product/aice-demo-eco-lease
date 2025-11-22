"""Export - Excel/PDF形式で見積書を出力"""

from pathlib import Path
from typing import Optional
from datetime import datetime
from loguru import logger

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.worksheet.pagebreak import Break

from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import Table, TableStyle
from reportlab.lib import colors

from pipelines.schemas import FMTDocument, EstimateItem, DisciplineType, ProjectInfo
from pipelines.pdf_generator import EcoleasePDFGenerator


class EstimateExporter:
    """見積書をExcel/PDF形式で出力"""

    def __init__(self, output_dir: str = "./output"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    @staticmethod
    def split_by_discipline_group(fmt_doc: FMTDocument) -> dict[str, FMTDocument]:
        """
        FMTDocumentを工事区分グループ別に分離

        電気・機械グループとガスグループに分ける

        Returns:
            {'electrical_mechanical': FMTDocument, 'gas': FMTDocument}
        """
        # グループ定義
        electrical_mechanical_group = {DisciplineType.ELECTRICAL, DisciplineType.MECHANICAL,
                                       DisciplineType.HVAC, DisciplineType.PLUMBING}
        gas_group = {DisciplineType.GAS}

        result = {}

        # 電気・機械グループ
        em_items = [item for item in fmt_doc.estimate_items
                    if item.discipline in electrical_mechanical_group or item.discipline is None]
        if em_items:
            em_disciplines = [d for d in fmt_doc.disciplines if d in electrical_mechanical_group]
            if em_disciplines:
                em_doc = FMTDocument(
                    fmt_version=fmt_doc.fmt_version,
                    created_at=fmt_doc.created_at,
                    project_info=fmt_doc.project_info.model_copy(),
                    facility_type=fmt_doc.facility_type,
                    building_specs=fmt_doc.building_specs,
                    disciplines=em_disciplines,
                    requirements=fmt_doc.requirements,
                    estimate_items=em_items,
                    legal_references=fmt_doc.legal_references,
                    qa_items=fmt_doc.qa_items,
                    raw_text=fmt_doc.raw_text,
                    extracted_tables=fmt_doc.extracted_tables,
                    metadata=fmt_doc.metadata.copy()
                )
                # 工事名に「電気・機械」を追加
                em_doc.project_info.project_name = f"{em_doc.project_info.project_name.replace('都市ガス設備工事', '').strip()} 電気・機械設備工事"
                result['electrical_mechanical'] = em_doc

        # ガスグループ
        gas_items = [item for item in fmt_doc.estimate_items
                     if item.discipline in gas_group]
        if gas_items:
            gas_disciplines = [d for d in fmt_doc.disciplines if d in gas_group]
            if gas_disciplines:
                gas_doc = FMTDocument(
                    fmt_version=fmt_doc.fmt_version,
                    created_at=fmt_doc.created_at,
                    project_info=fmt_doc.project_info.model_copy(),
                    facility_type=fmt_doc.facility_type,
                    building_specs=fmt_doc.building_specs,
                    disciplines=gas_disciplines,
                    requirements=fmt_doc.requirements,
                    estimate_items=gas_items,
                    legal_references=fmt_doc.legal_references,
                    qa_items=fmt_doc.qa_items,
                    raw_text=fmt_doc.raw_text,
                    extracted_tables=fmt_doc.extracted_tables,
                    metadata=fmt_doc.metadata.copy()
                )
                # 工事名を維持（都市ガス設備工事）
                result['gas'] = gas_doc

        return result

    def export_to_excel(self, fmt_doc: FMTDocument, filename: Optional[str] = None) -> str:
        """
        見積書をExcel形式で出力

        Args:
            fmt_doc: FMTドキュメント
            filename: 出力ファイル名（省略時は自動生成）

        Returns:
            出力ファイルパス
        """
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"見積書_{timestamp}.xlsx"

        output_path = self.output_dir / filename

        logger.info(f"Exporting estimate to Excel: {output_path}")

        # Workbook作成
        wb = Workbook()

        # シート1: 御見積書
        ws_summary = wb.active
        ws_summary.title = "御見積書"
        self._create_summary_sheet(ws_summary, fmt_doc)

        # シート2: 見積内訳明細書
        ws_detail = wb.create_sheet("見積内訳明細書")
        self._create_detail_sheet(ws_detail, fmt_doc)

        # 保存
        wb.save(output_path)

        logger.info(f"Excel file saved: {output_path}")

        return str(output_path)

    def _create_cover_sheet(self, ws, fmt_doc: FMTDocument):
        """送付状シートを作成（縦向き）"""

        # ページ設定を縦向きに
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
        ws.page_setup.paperSize = ws.PAPERSIZE_A4

        # タイトル
        ws['A1'] = "送　付　状"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:D1')

        # 日付
        ws['A3'] = datetime.now().strftime("%Y年%m月%d日")

        # 宛先
        ws['A5'] = f"{fmt_doc.project_info.client_name} 御中" if fmt_doc.project_info.client_name else "御中"
        ws['A5'].font = Font(size=12, bold=True)

        # 差出人
        ws['A7'] = "株式会社エコリース"
        ws['A8'] = "〒XXX-XXXX"
        ws['A9'] = "東京都XX区XXXX"

        # 件名
        ws['A11'] = "件名："
        ws['B11'] = fmt_doc.project_info.project_name
        ws['B11'].font = Font(bold=True)

        # 本文
        ws['A13'] = "拝啓　時下ますますご清栄のこととお慶び申し上げます。"
        ws['A14'] = "平素は格別のお引き立てを賜り、厚く御礼申し上げます。"
        ws['A16'] = "さて、下記の通り御見積書を提出いたします。"
        ws['A17'] = "ご査収の程、よろしくお願い申し上げます。"
        ws['A19'] = "敬具"

        # 記
        ws['A21'] = "記"
        ws['A21'].alignment = Alignment(horizontal='center')

        # 合計金額
        total_amount = sum(item.amount or 0 for item in fmt_doc.estimate_items if item.level == 0)
        ws['A23'] = "御見積金額："
        ws['B23'] = f"¥{total_amount:,.0f}（税別）"
        ws['B23'].font = Font(size=12, bold=True)

        ws['A25'] = "以上"
        ws['A25'].alignment = Alignment(horizontal='center')

    def _create_summary_sheet(self, ws, fmt_doc: FMTDocument):
        """御見積書（サマリー）シートを作成（縦向き、枠付き）"""

        # ページ設定を縦向きに
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
        ws.page_setup.paperSize = ws.PAPERSIZE_A4

        # 外枠を作成（A1:F30の範囲に枠線）
        thick_border = Border(
            left=Side(style='medium'),
            right=Side(style='medium'),
            top=Side(style='medium'),
            bottom=Side(style='medium')
        )

        # タイトル
        ws['A1'] = "御　見　積　書"
        ws['A1'].font = Font(size=18, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:F1')

        # 宛先
        ws['A3'] = f"{fmt_doc.project_info.client_name} 御中" if fmt_doc.project_info.client_name else "御中"
        ws['A3'].font = Font(size=12, bold=True)

        # 日付
        ws['E3'] = datetime.now().strftime("%Y年%m月%d日")

        # 差出人
        ws['E5'] = "株式会社エコリース"

        # 件名
        ws['A7'] = "件名："
        ws['B7'] = fmt_doc.project_info.project_name
        ws.merge_cells('B7:F7')

        # 合計金額
        total_amount = sum(item.amount or 0 for item in fmt_doc.estimate_items if item.level == 0)

        ws['A9'] = "御見積金額"
        ws['A9'].font = Font(size=14, bold=True)
        ws['A9'].alignment = Alignment(horizontal='center')

        ws['A10'] = f"¥{total_amount:,.0f}"
        ws['A10'].font = Font(size=16, bold=True)
        ws['A10'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A10:F10')

        ws['A11'] = "（消費税別途）"
        ws['A11'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A11:F11')

        # 内訳サマリー
        row = 13
        ws[f'A{row}'] = "内訳"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1

        # ヘッダー
        headers = ['No', '項目名', '金額']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row, col, header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')

        row += 1

        # 大項目のみ表示
        for item in fmt_doc.estimate_items:
            if item.level == 0:
                ws.cell(row, 1, item.item_no)
                ws.cell(row, 2, item.name)
                ws.cell(row, 3, f"¥{item.amount:,.0f}" if item.amount else "")
                row += 1

        # 合計
        ws.cell(row, 2, "合計").font = Font(bold=True)
        ws.cell(row, 3, f"¥{total_amount:,.0f}").font = Font(bold=True)

        # 御見積書全体に外枠をつける
        for row_num in range(1, row + 1):
            for col_num in range(1, 7):
                cell = ws.cell(row_num, col_num)
                if row_num == 1 or row_num == row:  # 上下端
                    if col_num == 1:  # 左上、左下
                        cell.border = Border(left=Side(style='medium'), top=Side(style='medium') if row_num == 1 else Side(style='medium'), bottom=Side(style='medium') if row_num == row else None)
                    elif col_num == 6:  # 右上、右下
                        cell.border = Border(right=Side(style='medium'), top=Side(style='medium') if row_num == 1 else Side(style='medium'), bottom=Side(style='medium') if row_num == row else None)
                    else:  # 上下端の中間
                        cell.border = Border(top=Side(style='medium') if row_num == 1 else Side(style='medium'), bottom=Side(style='medium') if row_num == row else None)
                else:  # 左右端
                    if col_num == 1:
                        cell.border = Border(left=Side(style='medium'))
                    elif col_num == 6:
                        cell.border = Border(right=Side(style='medium'))

    def _create_detail_sheet(self, ws, fmt_doc: FMTDocument):
        """見積内訳明細書シートを作成（横向き - Ecolease形式）"""

        # ページ設定を横向きに
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToHeight = False
        ws.page_setup.fitToWidth = 1

        # タイトル - 中央揃え
        ws['A1'] = "見　積　内　訳　明　細　書"
        ws['A1'].font = Font(name='MS Gothic', size=14, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:H1')

        # 見積番号
        ws['A2'] = f"({fmt_doc.metadata.get('quote_no', 'XXXXXXX-00')})"
        ws['A2'].font = Font(name='MS Gothic', size=9)

        # ヘッダー行
        row = 3
        headers = ['No', '名　　　称', '仕　　　様', '数　量', '単位', '単　　価', '金　　額', '摘　　要', '根拠情報']
        column_widths = [8, 30, 30, 10, 8, 15, 15, 20, 35]

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for col, (header, width) in enumerate(zip(headers, column_widths), start=1):
            cell = ws.cell(row, col, header)
            cell.font = Font(name='MS Gothic', size=9, bold=False)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='medium'),
                bottom=Side(style='medium')
            )
            # 列幅設定
            ws.column_dimensions[chr(64 + col)].width = width

        row += 1

        # データ行
        for item in fmt_doc.estimate_items:
            # インデント（階層に応じて）
            indent = "　" * item.level
            name = f"{indent}{item.name}"

            # No列 - 大項目のみ表示
            ws.cell(row, 1, item.item_no if not item.level else "")
            ws.cell(row, 1).alignment = Alignment(horizontal='center')
            ws.cell(row, 1).font = Font(name='MS Gothic', size=8)

            # 名称
            ws.cell(row, 2, name)
            ws.cell(row, 2).font = Font(name='MS Gothic', size=8, bold=(item.level == 0))

            # 仕様
            ws.cell(row, 3, item.specification or "")
            ws.cell(row, 3).font = Font(name='MS Gothic', size=8)

            # 数量
            qty_value = ""
            if item.quantity:
                qty_value = int(item.quantity) if item.quantity == int(item.quantity) else item.quantity
            ws.cell(row, 4, qty_value)
            ws.cell(row, 4).alignment = Alignment(horizontal='right')
            ws.cell(row, 4).font = Font(name='MS Gothic', size=8)

            # 単位
            ws.cell(row, 5, item.unit or "")
            ws.cell(row, 5).font = Font(name='MS Gothic', size=8)

            # 単価 - 詳細項目のみ表示
            unit_price_value = ""
            if item.unit_price and item.level > 0:
                unit_price_value = int(item.unit_price)
            ws.cell(row, 6, unit_price_value)
            ws.cell(row, 6).alignment = Alignment(horizontal='right')
            ws.cell(row, 6).font = Font(name='MS Gothic', size=8)
            if unit_price_value:
                ws.cell(row, 6).number_format = '#,##0'

            # 金額
            amount_value = ""
            if item.amount:
                amount_value = int(item.amount)
            ws.cell(row, 7, amount_value)
            ws.cell(row, 7).alignment = Alignment(horizontal='right')
            ws.cell(row, 7).font = Font(name='MS Gothic', size=8)
            if amount_value:
                ws.cell(row, 7).number_format = '#,##0'

            # 摘要
            ws.cell(row, 8, item.remarks or "")
            ws.cell(row, 8).font = Font(name='MS Gothic', size=8)

            # 根拠情報（新規追加）
            source_info = ""
            if item.source_reference:
                source_info = item.source_reference
            elif item.price_references:
                # KB IDsがある場合
                source_info = f"KB: {', '.join(item.price_references)}"

            ws.cell(row, 9, source_info)
            ws.cell(row, 9).font = Font(name='MS Gothic', size=7)
            ws.cell(row, 9).alignment = Alignment(wrap_text=True, vertical='top')

            # 罫線
            for col in range(1, 10):  # 9列に拡張
                ws.cell(row, col).border = thin_border

            row += 1

        # 総計行
        total_amount = sum(item.amount or 0 for item in fmt_doc.estimate_items if item.level == 0)
        ws.cell(row, 2, "総　　　計")
        ws.cell(row, 2).font = Font(name='MS Gothic', size=9, bold=True)
        ws.cell(row, 7, int(total_amount))
        ws.cell(row, 7).font = Font(name='MS Gothic', size=9, bold=True)
        ws.cell(row, 7).alignment = Alignment(horizontal='right')
        ws.cell(row, 7).number_format = '#,##0'

        # 総計行の罫線（上線を太く）
        for col in range(1, 10):  # 9列に拡張
            ws.cell(row, col).border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='medium'),
                bottom=Side(style='thin')
            )

        # フッター
        row += 2
        ws.cell(row, 1, "株式会社　　エコリース")
        ws.cell(row, 1).font = Font(name='MS Gothic', size=9)
        ws.cell(row, 8, "No　1")
        ws.cell(row, 8).font = Font(name='MS Gothic', size=9)
        ws.cell(row, 8).alignment = Alignment(horizontal='right')

    def export_to_pdf(self, fmt_doc: FMTDocument, filename: Optional[str] = None) -> str:
        """
        見積書をPDF形式で出力（Ecolease形式）

        Args:
            fmt_doc: FMTドキュメント
            filename: 出力ファイル名

        Returns:
            出力ファイルパス
        """
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"見積書_{timestamp}.pdf"

        output_path = self.output_dir / filename

        logger.info(f"Exporting estimate to PDF: {output_path}")

        # PDF生成
        pdf_gen = EcoleasePDFGenerator()
        pdf_gen.generate(fmt_doc, str(output_path))

        logger.info(f"PDF file saved: {output_path}")

        return str(output_path)

    def export_to_pdfs_by_discipline(self, fmt_doc: FMTDocument) -> list[str]:
        """
        見積書を工事区分グループ別に複数のPDFで出力

        Args:
            fmt_doc: FMTドキュメント

        Returns:
            出力ファイルパスのリスト
        """
        logger.info("Splitting document by discipline groups...")

        # 分野別に分離
        docs_by_group = self.split_by_discipline_group(fmt_doc)

        if not docs_by_group:
            logger.warning("No discipline groups found, exporting as single PDF")
            return [self.export_to_pdf(fmt_doc)]

        output_paths = []
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        for group_name, group_doc in docs_by_group.items():
            # ファイル名を分野別に生成
            if group_name == 'electrical_mechanical':
                filename = f"見積書_電気機械_{timestamp}.pdf"
            elif group_name == 'gas':
                filename = f"見積書_ガス_{timestamp}.pdf"
            else:
                filename = f"見積書_{group_name}_{timestamp}.pdf"

            path = self.export_to_pdf(group_doc, filename)
            output_paths.append(path)
            logger.info(f"Generated PDF for {group_name}: {path}")

        return output_paths

    def export_to_pdf_old(self, fmt_doc: FMTDocument, filename: Optional[str] = None) -> str:
        """
        見積書をPDF形式で出力

        Args:
            fmt_doc: FMTドキュメント
            filename: 出力ファイル名

        Returns:
            出力ファイルパス
        """
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"見積書_{timestamp}.pdf"

        output_path = self.output_dir / filename

        logger.info(f"Exporting estimate to PDF: {output_path}")

        # 日本語フォント登録（システムフォントを使用）
        try:
            # macOSの場合
            pdfmetrics.registerFont(TTFont('Japanese', '/System/Library/Fonts/ヒラギノ角ゴシック W3.ttc'))
            font_name = 'Japanese'
        except:
            try:
                # Linuxの場合（IPAフォント）
                pdfmetrics.registerFont(TTFont('Japanese', '/usr/share/fonts/opentype/ipaexfont-gothic/ipaexg.ttf'))
                font_name = 'Japanese'
            except:
                # フォントが見つからない場合はHelveticaを使用（日本語は表示されない）
                logger.warning("Japanese font not found, using Helvetica")
                font_name = 'Helvetica'

        # PDFキャンバス作成
        c = canvas.Canvas(str(output_path), pagesize=A4)
        width, height = A4

        # 1ページ目：御見積書（サマリー）
        self._create_pdf_summary_page(c, fmt_doc, width, height, font_name)
        c.showPage()

        # 2ページ目以降：見積内訳明細書（横向き）
        c.setPageSize(landscape(A4))
        lwidth, lheight = landscape(A4)
        self._create_pdf_detail_pages(c, fmt_doc, lwidth, lheight, font_name)

        # PDF保存
        c.save()

        logger.info(f"PDF file saved: {output_path}")

        return str(output_path)

    def _create_pdf_summary_page(self, c, fmt_doc: FMTDocument, width, height, font_name):
        """御見積書サマリーページを作成"""

        # タイトル
        c.setFont(font_name, 18)
        c.drawCentredString(width / 2, height - 50*mm, "御　見　積　書")

        # 宛先
        c.setFont(font_name, 12)
        client_name = fmt_doc.project_info.client_name or ""
        c.drawString(30*mm, height - 70*mm, f"{client_name} 御中")

        # 日付
        c.setFont(font_name, 10)
        c.drawRightString(width - 30*mm, height - 70*mm, datetime.now().strftime("%Y年%m月%d日"))

        # 差出人
        c.drawRightString(width - 30*mm, height - 80*mm, "株式会社エコリース")

        # 件名
        c.setFont(font_name, 10)
        c.drawString(30*mm, height - 100*mm, "件名：")
        c.drawString(50*mm, height - 100*mm, fmt_doc.project_info.project_name)

        # 合計金額
        total_amount = sum(item.amount or 0 for item in fmt_doc.estimate_items if item.level == 0)
        c.setFont(font_name, 14)
        c.drawString(30*mm, height - 120*mm, "御見積金額")
        c.setFont(font_name, 16)
        c.drawCentredString(width / 2, height - 135*mm, f"¥{total_amount:,.0f}")
        c.setFont(font_name, 10)
        c.drawCentredString(width / 2, height - 145*mm, "（消費税別途）")

        # 内訳サマリー
        c.setFont(font_name, 12)
        c.drawString(30*mm, height - 165*mm, "内訳")

        y_position = height - 180*mm
        c.setFont(font_name, 10)

        for item in fmt_doc.estimate_items:
            if item.level == 0:
                c.drawString(35*mm, y_position, f"{item.item_no}. {item.name}")
                c.drawRightString(width - 35*mm, y_position, f"¥{item.amount:,.0f}" if item.amount else "")
                y_position -= 5*mm

        # 合計
        y_position -= 5*mm
        c.setFont(font_name, 12)
        c.drawString(35*mm, y_position, "合計")
        c.drawRightString(width - 35*mm, y_position, f"¥{total_amount:,.0f}")

    def _create_pdf_detail_pages(self, c, fmt_doc: FMTDocument, width, height, font_name):
        """見積内訳明細書ページを作成（横向き）"""

        # タイトル
        c.setFont(font_name, 14)
        c.drawCentredString(width / 2, height - 20*mm, "見積内訳明細書")

        # 案件名
        c.setFont(font_name, 9)
        c.drawString(20*mm, height - 30*mm, f"案件名：{fmt_doc.project_info.project_name}")

        # テーブルデータを準備
        data = [['No', '名称', '仕様', '数量', '単位', '単価', '金額', '摘要']]

        for item in fmt_doc.estimate_items:
            indent = "　" * item.level
            row = [
                item.item_no,
                f"{indent}{item.name}",
                item.specification or "",
                str(item.quantity) if item.quantity else "",
                item.unit or "",
                f"¥{item.unit_price:,.0f}" if item.unit_price else "",
                f"¥{item.amount:,.0f}" if item.amount else "",
                item.remarks or ""
            ]
            data.append(row)

        # 合計行
        total_amount = sum(item.amount or 0 for item in fmt_doc.estimate_items if item.level == 0)
        data.append(['', '', '', '', '', '合計', f"¥{total_amount:,.0f}", ''])

        # テーブル作成（列幅を調整）
        col_widths = [15*mm, 50*mm, 50*mm, 15*mm, 12*mm, 25*mm, 25*mm, 30*mm]

        # ページに収まるように行を分割
        rows_per_page = 25
        for page_start in range(0, len(data), rows_per_page):
            if page_start > 0:
                c.showPage()
                c.setPageSize(landscape(A4))
                c.setFont(font_name, 14)
                c.drawCentredString(width / 2, height - 20*mm, "見積内訳明細書（続き）")

            page_data = data[page_start:page_start + rows_per_page]
            if page_start > 0:
                # 2ページ目以降はヘッダーを再度追加
                page_data = [data[0]] + page_data

            table = Table(page_data, colWidths=col_widths)
            table.setStyle(TableStyle([
                ('FONT', (0, 0), (-1, -1), font_name, 8),
                ('FONT', (0, 0), (-1, 0), font_name, 9),  # ヘッダーは少し大きく
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('ALIGN', (3, 0), (3, -1), 'RIGHT'),  # 数量
                ('ALIGN', (5, 0), (6, -1), 'RIGHT'),  # 単価・金額
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                ('FONTNAME', (0, -1), (-1, -1), font_name),  # 合計行
                ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
            ]))

            # テーブルを描画
            table.wrapOn(c, width, height)
            table.drawOn(c, 15*mm, height - 40*mm - len(page_data) * 6*mm)
