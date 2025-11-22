"""過去見積KBとLLM抽出機能の統合モジュール（拡張版）"""

import os
import json
import re
from pathlib import Path
from typing import List, Dict, Any, Optional, Union
from datetime import datetime, date
from collections import defaultdict
import statistics
from dotenv import load_dotenv
from anthropic import Anthropic
from loguru import logger
import PyPDF2
import openpyxl

from pipelines.schemas import (
    PriceReference, DisciplineType, EstimateItem,
    Requirement, LegalReference
)
from pipelines.cost_tracker import record_cost


class PriceKBBuilder:
    """
    見積書から過去見積KBを構築（Excel/PDF対応、複数案件統合機能付き）

    機能:
    - PDF見積書からのKB化（テキスト/OCR）
    - Excel見積書からのKB化
    - 複数案件の価格統合（平均/中央値/時系列重み）
    - 既存KBとのマージ
    """

    def __init__(self, kb_path: str = "kb/price_kb.json"):
        load_dotenv()
        self.client = Anthropic(api_key=os.getenv("ANTHROPIC_API_KEY"))
        self.model_name = os.getenv("CLAUDE_MODEL", "claude-sonnet-4-20250514")
        self.kb_path = kb_path
        self.kb_items: List[Dict[str, Any]] = []

        # 既存KBを読み込み
        if Path(kb_path).exists():
            with open(kb_path, 'r', encoding='utf-8') as f:
                self.kb_items = json.load(f)
            logger.info(f"Loaded {len(self.kb_items)} items from KB")
        else:
            logger.info(f"No existing KB found at {kb_path}, starting fresh")

    def extract_estimate_from_pdf(self, pdf_path: str) -> List[PriceReference]:
        """見積書PDFから価格情報を抽出してKB化（OCR対応）"""
        logger.info(f"Building price KB from: {pdf_path}")

        # PDFからテキストを抽出（全ページ対応）
        with open(pdf_path, 'rb') as file:
            pdf_reader = PyPDF2.PdfReader(file)
            text = ""
            total_pages = len(pdf_reader.pages)
            # 全ページを処理（制限なし）
            for page_num in range(total_pages):
                text += pdf_reader.pages[page_num].extract_text() + "\n"

        logger.info(f"Extracted {len(text)} characters from PDF ({total_pages} pages)")

        # テキストがほとんど抽出できない場合はOCRを使用
        # 閾値を緩和: 100 → 500文字（スキャンPDFの検出精度向上）
        if len(text.strip()) < 500:
            logger.warning("Text extraction failed, using OCR...")
            from pipelines.ocr_extractor import OCRExtractor
            ocr = OCRExtractor()
            items_data = ocr.extract_from_pdf(pdf_path)

            # KB形式に変換
            price_refs = []
            project_name = Path(pdf_path).stem

            for i, item in enumerate(items_data):
                if item.get("unit_price") and item.get("unit_price") > 0:
                    # 工事区分のマッピング
                    discipline_map = {
                        "電気": DisciplineType.ELECTRICAL,
                        "機械": DisciplineType.MECHANICAL,
                        "空調": DisciplineType.HVAC,
                        "衛生": DisciplineType.PLUMBING,
                        "ガス": DisciplineType.GAS,
                        "消防": DisciplineType.FIRE_PROTECTION
                    }

                    # OCR結果の工事区分を使用、なければキーワードベースで推定
                    ocr_discipline = item.get("discipline", "").replace("設備工事", "")
                    if ocr_discipline in discipline_map:
                        discipline = discipline_map[ocr_discipline]
                    else:
                        # キーワードベースで自動推定（ファイル名もヒント）
                        discipline = self._infer_discipline(
                            item.get("name", ""),
                            item.get("specification", ""),
                            project_name
                        )

                    # コンテキストタグの生成
                    context_tags = []
                    if "学校" in project_name or "高校" in project_name:
                        context_tags.append("学校")
                    if "改修" in project_name:
                        context_tags.append("改修")
                    if "仮設" in project_name:
                        context_tags.append("仮設")

                    price_ref = PriceReference(
                        item_id=f"{project_name}_{i+1:03d}",
                        description=item.get("name", ""),
                        discipline=discipline,
                        unit=item.get("unit", "式"),
                        unit_price=float(item.get("unit_price", 0)),
                        vendor=None,
                        valid_from=date.today(),
                        valid_to=None,
                        source_project=project_name,
                        context_tags=context_tags,
                        features={
                            "specification": item.get("specification", ""),
                            "quantity": item.get("quantity"),
                        },
                        similarity_score=0.0
                    )
                    price_refs.append(price_ref)

            logger.info(f"Extracted {len(price_refs)} items using OCR")
            return price_refs

        # LLMで構造化データに変換
        # テキスト制限を緩和: 最大60000文字（大規模PDF対応）
        prompt = f"""以下の見積書PDFから、単価情報を抽出して単価データベース（KB）を構築します。

見積書テキスト:
{text[:60000]}

【抽出する情報】
各見積項目について：
1. 項目名（name）- 具体的で検索可能な名称
2. 仕様（specification）- サイズ、型番、材質等
3. 数量（quantity）
4. 単位（unit）
5. 単価（unit_price）- 必須
6. 金額（amount）
7. 工事区分（discipline）- 下記から選択

【工事区分の判定基準】
- 電気: キュービクル、分電盤、配電盤、ケーブル、CV、配線、照明、コンセント、接地、ブレーカー
- 機械: ダクト、換気扇、ファン、ポンプ、エレベーター、ボイラー
- 空調: エアコン、パッケージ、室外機、室内機、冷媒配管、ヒートポンプ
- 衛生: 給水管、排水管、給湯、トイレ、洗面、受水槽
- ガス: ガス管、ガスコンセント、ガス栓、PE管、ガスメーター
- 消防: スプリンクラー、感知器、消火栓、誘導灯、火災報知

【重要な処理ルール】
1. **「同上」「〃」「上記と同じ」などの省略表記は、直前の具体的な項目名に置き換えて出力してください**
   例: 「同上施工費」→「架橋ポリエチレンケーブル施工費」
   例: 「同上支持材」→「600Vビニル絶縁電線支持材」

2. **以下の項目は除外してください:**
   - 小計・合計・計行
   - 項目名が空または「計」「小計」「合計」のみのもの
   - 単価が0円または記載なしのもの
   - 諸経費、法定福利費（率計算のため）

3. **項目名は単独で意味が通じる形にしてください:**
   悪い例: 「同上」「〃」「設置費」
   良い例: 「気中開閉器設置費」「CVTケーブル布設費」

【出力形式】
JSON配列で出力してください：
```json
[
  {{
    "name": "架橋ポリエチレンケーブル",
    "specification": "6KV CVT38sq",
    "quantity": 153,
    "unit": "m",
    "unit_price": 10300,
    "amount": 1575900,
    "discipline": "電気"
  }},
  {{
    "name": "架橋ポリエチレンケーブル布設費",
    "specification": "6KV CVT38sq",
    "quantity": 153,
    "unit": "m",
    "unit_price": 2500,
    "amount": 382500,
    "discipline": "電気"
  }}
]
```"""

        try:
            response = self.client.messages.create(
                model=self.model_name,
                max_tokens=16000,  # 詳細な抽出のため増加
                temperature=0,
                messages=[{"role": "user", "content": prompt}]
            )

            # コスト記録
            record_cost(
                operation="KB抽出（単価）",
                model_name=self.model_name,
                input_tokens=response.usage.input_tokens,
                output_tokens=response.usage.output_tokens,
                metadata={"file": Path(pdf_path).name}
            )

            response_text = response.content[0].text

            # JSONを抽出
            json_start = response_text.find('[')
            json_end = response_text.rfind(']') + 1

            if json_start == -1 or json_end == 0:
                logger.error("No JSON found in response")
                return []

            json_str = response_text[json_start:json_end]
            items_data = json.loads(json_str)

            logger.info(f"Extracted {len(items_data)} price items")

            # PriceReferenceオブジェクトに変換
            price_refs = []
            project_name = Path(pdf_path).stem

            for i, item in enumerate(items_data):
                if item.get("unit_price") and item.get("unit_price") > 0:
                    # 工事区分のマッピング
                    discipline_map = {
                        "電気": DisciplineType.ELECTRICAL,
                        "機械": DisciplineType.MECHANICAL,
                        "空調": DisciplineType.HVAC,
                        "衛生": DisciplineType.PLUMBING,
                        "ガス": DisciplineType.GAS,
                        "消防": DisciplineType.FIRE_PROTECTION
                    }

                    # LLMの結果を使用、なければキーワードベースで推定
                    llm_discipline = item.get("discipline", "")
                    if llm_discipline in discipline_map:
                        discipline = discipline_map[llm_discipline]
                    else:
                        # キーワードベースで自動推定（ファイル名もヒント）
                        discipline = self._infer_discipline(
                            item.get("name", ""),
                            item.get("specification", ""),
                            project_name
                        )

                    # コンテキストタグの生成
                    context_tags = []
                    if "学校" in project_name or "高校" in project_name:
                        context_tags.append("学校")
                    if "改修" in project_name:
                        context_tags.append("改修")
                    if "仮設" in project_name:
                        context_tags.append("仮設")

                    price_ref = PriceReference(
                        item_id=f"{project_name}_{i+1:03d}",
                        description=item.get("name", ""),
                        discipline=discipline,
                        unit=item.get("unit", "式"),
                        unit_price=float(item.get("unit_price", 0)),
                        vendor=None,
                        valid_from=date.today(),
                        valid_to=None,
                        source_project=project_name,
                        context_tags=context_tags,
                        features={
                            "specification": item.get("specification", ""),
                            "quantity": item.get("quantity"),
                        },
                        similarity_score=0.0
                    )
                    price_refs.append(price_ref)

            return price_refs

        except Exception as e:
            logger.error(f"Error extracting prices: {e}")
            return []

    def save_kb_to_json(self, price_refs: List[PriceReference], output_path: str):
        """KBをJSONファイルに保存"""
        kb_data = [ref.model_dump(mode='json') for ref in price_refs]

        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(kb_data, f, ensure_ascii=False, indent=2, default=str)

        logger.info(f"Saved {len(price_refs)} price references to {output_path}")

    def load_kb_from_json(self, kb_path: str) -> List[PriceReference]:
        """JSONファイルからKBを読み込み（古いフォーマット対応）"""
        with open(kb_path, 'r', encoding='utf-8') as f:
            kb_data = json.load(f)

        price_refs = []
        for item in kb_data:
            # 古いKBフォーマットに対応（必須フィールドのデフォルト値）
            if 'valid_from' not in item:
                item['valid_from'] = date.today()
            if 'source_project' not in item:
                item['source_project'] = item.get('item_id', 'unknown').split('_')[0]
            if 'valid_to' not in item:
                item['valid_to'] = None
            if 'vendor' not in item:
                item['vendor'] = None
            if 'similarity_score' not in item:
                item['similarity_score'] = 0.0
            if 'context_tags' not in item:
                item['context_tags'] = []

            price_refs.append(PriceReference(**item))

        logger.info(f"Loaded {len(price_refs)} price references from {kb_path}")
        return price_refs

    def extract_estimate_from_excel(self, excel_path: str, project_name: str = None) -> List[PriceReference]:
        """Excel見積書から価格情報を抽出してKB化

        Args:
            excel_path: Excel見積書のパス
            project_name: プロジェクト名（指定しない場合はファイル名）

        Returns:
            PriceReferenceのリスト
        """
        logger.info(f"Building price KB from Excel: {excel_path}")

        if project_name is None:
            project_name = Path(excel_path).stem

        try:
            wb = openpyxl.load_workbook(excel_path, data_only=True)
            ws = wb.active

            price_refs = []
            item_counter = 1

            # ヘッダー行を探索（「名称」「単価」などを含む行）
            header_row = None
            for row_idx in range(1, min(20, ws.max_row + 1)):
                row_values = [str(cell.value or "").strip() for cell in ws[row_idx]]
                if any("名称" in v or "名　　称" in v for v in row_values):
                    header_row = row_idx
                    logger.info(f"Found header at row {header_row}")
                    break

            if not header_row:
                logger.error("Could not find header row in Excel")
                return []

            # 列インデックスを特定
            header_cells = [str(cell.value or "").strip() for cell in ws[header_row]]

            name_col = None
            spec_col = None
            quantity_col = None
            unit_col = None
            unit_price_col = None

            for idx, header in enumerate(header_cells):
                if "名称" in header or "名　　称" in header:
                    name_col = idx
                elif "仕様" in header or "仕　　様" in header:
                    spec_col = idx
                elif "数量" in header or "数　量" in header:
                    quantity_col = idx
                elif "単位" in header:
                    unit_col = idx
                elif "単価" in header or "単　　価" in header:
                    unit_price_col = idx

            logger.info(f"Column mapping: name={name_col}, spec={spec_col}, qty={quantity_col}, unit={unit_col}, price={unit_price_col}")

            if name_col is None or unit_price_col is None:
                logger.error("Required columns (name, unit_price) not found")
                return []

            # データ行を処理
            for row_idx in range(header_row + 1, ws.max_row + 1):
                row = ws[row_idx]

                name = str(row[name_col].value or "").strip() if name_col is not None else ""
                spec = str(row[spec_col].value or "").strip() if spec_col is not None else ""
                quantity = row[quantity_col].value if quantity_col is not None else None
                unit = str(row[unit_col].value or "").strip() if unit_col is not None else "式"
                unit_price = row[unit_price_col].value if unit_price_col is not None else None

                # 空行・親項目（単価なし）をスキップ
                if not name or not unit_price or unit_price <= 0:
                    continue

                # 工事区分を推定（ファイル名もヒントとして使用）
                discipline = self._infer_discipline(name, spec, project_name)

                # コンテキストタグ生成
                context_tags = []
                if "学校" in project_name or "高校" in project_name:
                    context_tags.append("学校")
                if "改修" in project_name:
                    context_tags.append("改修")
                if "仮設" in project_name:
                    context_tags.append("仮設")

                price_ref = PriceReference(
                    item_id=f"{project_name}_{item_counter:03d}",
                    description=name,
                    discipline=discipline,
                    unit=unit,
                    unit_price=float(unit_price),
                    vendor=None,
                    valid_from=date.today(),
                    valid_to=None,
                    source_project=project_name,
                    context_tags=context_tags,
                    features={
                        "specification": spec,
                        "quantity": quantity,
                    },
                    similarity_score=0.0
                )
                price_refs.append(price_ref)
                item_counter += 1

            logger.info(f"Extracted {len(price_refs)} price items from Excel")
            return price_refs

        except Exception as e:
            logger.error(f"Error extracting from Excel: {e}")
            return []

    def _infer_discipline(self, name: str, spec: str, filename: str = "") -> DisciplineType:
        """項目名・仕様・ファイル名から工事区分を推定

        Args:
            name: 項目名
            spec: 仕様
            filename: 元ファイル名（追加のヒントとして使用）

        Returns:
            推定された工事区分
        """
        text = (name + " " + spec).lower()
        filename_lower = filename.lower()

        # ガス設備工事のキーワード（最も具体的）
        gas_keywords = [
            "ガス", "都市ガス", "lpガス", "プロパン",
            "ガス管", "白ガス管", "pe管", "ガスコンセント",
            "ネジコック", "分岐コック", "ガス栓", "ガスメーター",
            "ガス漏れ", "遮断弁", "ガス工事"
        ]

        # 電気設備工事のキーワード
        electrical_keywords = [
            "電気", "電線", "ケーブル", "cv", "cvt", "iv",
            "配線", "幹線", "分電盤", "配電盤", "動力盤",
            "コンセント", "照明", "led", "蛍光灯", "器具",
            "スイッチ", "接地", "アース", "ブレーカー", "遮断器",
            "キュービクル", "受電", "変電", "変圧器", "トランス",
            "pas", "高圧", "低圧", "電灯", "動力",
            "インターホン", "放送", "lan", "弱電", "通信",
            "火災報知", "自火報", "非常照明", "誘導灯",
            "避雷", "雷", "接地工事", "電気工事"
        ]

        # 機械設備工事のキーワード
        mechanical_keywords = [
            "機械", "ダクト", "換気", "送風", "排風",
            "ファン", "モーター", "ポンプ", "揚水",
            "エレベーター", "昇降機", "リフト",
            "ボイラー", "熱源", "チラー",
            "機械工事", "機械設備"
        ]

        # 空調設備工事のキーワード
        hvac_keywords = [
            "空調", "エアコン", "冷暖房", "冷房", "暖房",
            "パッケージ", "ヒートポンプ", "室外機", "室内機",
            "ビル用マルチ", "ghp", "ehp",
            "冷媒", "r410", "r32", "フロン",
            "全熱交換", "ロスナイ", "換気空調",
            "空調工事", "空調設備"
        ]

        # 衛生設備工事のキーワード
        plumbing_keywords = [
            "衛生", "給水", "排水", "給湯", "温水",
            "配水管", "給水管", "排水管", "汚水", "雑排水",
            "トイレ", "便器", "洗面", "流し", "シンク",
            "受水槽", "高架水槽", "貯水槽",
            "衛生器具", "水栓", "バルブ",
            "浄化槽", "グリストラップ",
            "衛生工事", "衛生設備", "給排水"
        ]

        # 消防設備工事のキーワード
        fire_keywords = [
            "消防", "消火", "スプリンクラー", "消火栓",
            "消火器", "泡消火", "ガス消火",
            "火災報知", "感知器", "発信機", "警報",
            "避難", "誘導灯", "非常放送",
            "防火", "防排煙", "排煙",
            "消防工事", "消防設備"
        ]

        # ファイル名からのヒント（優先度高）
        if "ガス" in filename_lower or "gas" in filename_lower:
            # ただし電気・機械が含まれていなければ
            if "電気" not in filename_lower and "機械" not in filename_lower:
                return DisciplineType.GAS
        if "電気" in filename_lower or "electrical" in filename_lower:
            return DisciplineType.ELECTRICAL
        if "機械" in filename_lower or "mechanical" in filename_lower:
            return DisciplineType.MECHANICAL
        if "空調" in filename_lower or "hvac" in filename_lower:
            return DisciplineType.HVAC
        if "衛生" in filename_lower or "給排水" in filename_lower:
            return DisciplineType.PLUMBING
        if "消防" in filename_lower or "fire" in filename_lower:
            return DisciplineType.FIRE_PROTECTION

        # 項目名・仕様からの判定（キーワードマッチング）
        # スコアリング方式で最も該当するものを選択
        scores = {
            DisciplineType.GAS: sum(1 for kw in gas_keywords if kw in text),
            DisciplineType.ELECTRICAL: sum(1 for kw in electrical_keywords if kw in text),
            DisciplineType.MECHANICAL: sum(1 for kw in mechanical_keywords if kw in text),
            DisciplineType.HVAC: sum(1 for kw in hvac_keywords if kw in text),
            DisciplineType.PLUMBING: sum(1 for kw in plumbing_keywords if kw in text),
            DisciplineType.FIRE_PROTECTION: sum(1 for kw in fire_keywords if kw in text),
        }

        max_score = max(scores.values())
        if max_score > 0:
            for discipline, score in scores.items():
                if score == max_score:
                    return discipline

        # どれにも該当しない場合は「その他」としてMECHANICALを返す
        return DisciplineType.MECHANICAL

    def aggregate_multiple_estimates(
        self,
        estimate_paths: List[str],
        method: str = "median"
    ) -> List[PriceReference]:
        """複数見積から価格を統合

        Args:
            estimate_paths: 見積ファイルパスのリスト（Excel/PDF）
            method: 統合方法 ("median" | "average" | "time_weighted")

        Returns:
            統合されたPriceReferenceのリスト
        """
        logger.info(f"Aggregating {len(estimate_paths)} estimates using {method} method")

        all_refs: List[PriceReference] = []

        # 各ファイルからデータ抽出
        for path in estimate_paths:
            if path.endswith('.xlsx') or path.endswith('.xls'):
                refs = self.extract_estimate_from_excel(path)
            elif path.endswith('.pdf'):
                refs = self.extract_estimate_from_pdf(path)
            else:
                logger.warning(f"Unsupported file type: {path}")
                continue

            all_refs.extend(refs)

        logger.info(f"Total items before aggregation: {len(all_refs)}")

        # 同一項目をグループ化（description, specification, unitで）
        grouped: Dict[tuple, List[PriceReference]] = defaultdict(list)

        for ref in all_refs:
            key = (
                ref.description,
                ref.features.get("specification", ""),
                ref.unit
            )
            grouped[key].append(ref)

        # 価格を統合
        aggregated_refs = []

        for key, refs in grouped.items():
            description, specification, unit = key

            if len(refs) == 1:
                # 単一データの場合はそのまま
                aggregated_refs.append(refs[0])
                continue

            # 価格リスト
            prices = [ref.unit_price for ref in refs]

            # 統合方法に応じて価格を計算
            if method == "median":
                aggregated_price = statistics.median(prices)
            elif method == "average":
                aggregated_price = statistics.mean(prices)
            elif method == "time_weighted":
                # 新しい見積ほど重みを大きく（簡易実装）
                weights = list(range(1, len(refs) + 1))
                aggregated_price = sum(p * w for p, w in zip(prices, weights)) / sum(weights)
            else:
                aggregated_price = statistics.median(prices)  # デフォルト

            # 統合されたPriceReferenceを作成
            source_projects = list(set(ref.source_project for ref in refs))

            aggregated_ref = PriceReference(
                item_id=f"AGG_{refs[0].item_id}",
                description=description,
                discipline=refs[0].discipline,
                unit=unit,
                unit_price=aggregated_price,
                vendor=None,
                valid_from=min(ref.valid_from for ref in refs),
                valid_to=None,
                source_project=", ".join(source_projects),
                context_tags=list(set(tag for ref in refs for tag in ref.context_tags)),
                features={
                    "specification": specification,
                    "aggregated_from": len(refs),
                    "price_range": f"¥{min(prices):,.0f} - ¥{max(prices):,.0f}",
                    "std_dev": statistics.stdev(prices) if len(prices) > 1 else 0,
                },
                similarity_score=0.0
            )
            aggregated_refs.append(aggregated_ref)

        logger.info(f"Aggregated to {len(aggregated_refs)} unique items")
        return aggregated_refs

    def merge_with_existing_kb(
        self,
        new_refs: List[PriceReference],
        merge_strategy: str = "keep_new"
    ) -> List[PriceReference]:
        """新しいKBエントリを既存KBとマージ

        Args:
            new_refs: 新しいPriceReferenceリスト
            merge_strategy: マージ戦略 ("keep_new" | "keep_old" | "average")

        Returns:
            マージされたPriceReferenceリスト
        """
        logger.info(f"Merging {len(new_refs)} new items with existing KB ({len(self.kb_items)} items)")

        # 既存KBをPriceReferenceに変換
        existing_refs = self.load_kb_from_json(self.kb_path) if Path(self.kb_path).exists() else []

        # 既存項目をマッピング
        existing_map: Dict[tuple, PriceReference] = {}
        for ref in existing_refs:
            key = (
                ref.description,
                ref.features.get("specification", ""),
                ref.unit
            )
            existing_map[key] = ref

        merged_refs = []
        added_count = 0
        updated_count = 0

        for new_ref in new_refs:
            key = (
                new_ref.description,
                new_ref.features.get("specification", ""),
                new_ref.unit
            )

            if key in existing_map:
                # 既存項目がある場合
                existing_ref = existing_map[key]

                if merge_strategy == "keep_new":
                    merged_refs.append(new_ref)
                elif merge_strategy == "keep_old":
                    merged_refs.append(existing_ref)
                elif merge_strategy == "average":
                    # 価格を平均
                    avg_price = (existing_ref.unit_price + new_ref.unit_price) / 2
                    merged_ref = existing_ref.model_copy(update={"unit_price": avg_price})
                    merged_refs.append(merged_ref)

                updated_count += 1
                # マップから削除（処理済み）
                del existing_map[key]
            else:
                # 新規項目
                merged_refs.append(new_ref)
                added_count += 1

        # 残った既存項目を追加
        merged_refs.extend(existing_map.values())

        logger.info(f"Merge complete: {added_count} added, {updated_count} updated, {len(merged_refs)} total")
        return merged_refs


class EnhancedEstimateExtractor:
    """EstimateExtractorに信頼度スコアと根拠情報を追加"""

    def __init__(self, price_kb: List[PriceReference]):
        load_dotenv()
        self.client = Anthropic(api_key=os.getenv("ANTHROPIC_API_KEY"))
        self.model_name = os.getenv("CLAUDE_MODEL", "claude-sonnet-4-20250514")
        self.price_kb = price_kb
        logger.info(f"Initialized with {len(price_kb)} price references")

    def extract_with_confidence(self, spec_text: str, discipline: DisciplineType) -> List[EstimateItem]:
        """
        信頼度スコア付きで見積項目を抽出

        Returns:
            信頼度スコアと根拠情報を含むEstimateItemのリスト
        """
        logger.info(f"Extracting items with confidence for {discipline}")

        discipline_map = {
            DisciplineType.ELECTRICAL: "電気設備工事",
            DisciplineType.MECHANICAL: "機械設備工事",
            DisciplineType.HVAC: "空調設備工事",
            DisciplineType.PLUMBING: "衛生設備工事",
            DisciplineType.GAS: "都市ガス設備工事",
            DisciplineType.FIRE_PROTECTION: "消防設備工事"
        }

        discipline_name = discipline_map.get(discipline, "設備工事")

        prompt = f"""以下の入札仕様書から、{discipline_name}に関連する見積項目を抽出してください。

仕様書テキスト:
{spec_text[:15000]}

【抽出する項目】
各見積項目について、以下の情報を抽出してください：
1. 項目名（name）
2. 仕様（specification）
3. 数量（quantity）: 明記されている場合のみ
4. 単位（unit）
5. 階層レベル（level）: 0=親項目, 1=子項目, 2=孫項目
6. 信頼度（confidence）: 0.0-1.0（仕様書に明記=0.9-1.0、推測=0.3-0.6、不明=0.0-0.2）
7. 根拠（source_page）: 該当ページ番号があれば

【重要】
- 数量が明記されていない場合は null を設定
- 単価は設定しない（後でKBから検索）
- 信頼度は以下の基準で設定:
  * 1.0: 仕様書に具体的な数値で明記
  * 0.8: 仕様書に記載あるが曖昧
  * 0.5: 図面や文脈から推測可能
  * 0.3: 一般的な標準値
  * 0.0: 不明・要確認

【出力形式】
JSON配列形式で出力してください：
```json
[
  {{
    "item_no": "1",
    "level": 0,
    "name": "{discipline_name}",
    "specification": "",
    "quantity": null,
    "unit": "式",
    "confidence": 0.9,
    "source_page": 1
  }},
  {{
    "item_no": "",
    "level": 2,
    "name": "白ガス管（ネジ接合）",
    "specification": "15A",
    "quantity": 93,
    "unit": "m",
    "confidence": 1.0,
    "source_page": 5
  }}
]
```"""

        try:
            response = self.client.messages.create(
                model=self.model_name,
                max_tokens=8000,
                temperature=0,
                messages=[{"role": "user", "content": prompt}]
            )

            # コスト記録
            record_cost(
                operation="見積抽出（信頼度付き）",
                model_name=self.model_name,
                input_tokens=response.usage.input_tokens,
                output_tokens=response.usage.output_tokens,
                metadata={"discipline": discipline.value}
            )

            response_text = response.content[0].text

            # JSONを抽出
            json_start = response_text.find('[')
            json_end = response_text.rfind(']') + 1

            if json_start == -1 or json_end == 0:
                logger.error("No JSON found in response")
                return []

            json_str = response_text[json_start:json_end]
            items_data = json.loads(json_str)

            # EstimateItemオブジェクトに変換
            estimate_items = []
            for item_data in items_data:
                estimate_item = EstimateItem(
                    item_no=item_data.get("item_no", ""),
                    level=item_data.get("level", 0),
                    name=item_data.get("name", ""),
                    specification=item_data.get("specification", ""),
                    quantity=item_data.get("quantity"),
                    unit=item_data.get("unit", ""),
                    unit_price=None,
                    amount=None,
                    discipline=discipline,
                    confidence=item_data.get("confidence", 0.5),
                    source_type="llm_extraction",
                    source_reference=f"spec_page_{item_data.get('source_page', 'unknown')}"
                )
                estimate_items.append(estimate_item)

            logger.info(f"Extracted {len(estimate_items)} items with confidence scores")
            return estimate_items

        except Exception as e:
            logger.error(f"Error extracting items with confidence: {e}")
            return []

    def enrich_with_price_rag(self, items: List[EstimateItem]) -> List[EstimateItem]:
        """
        KBから単価を検索して付与（簡易RAG）

        Args:
            items: 単価未設定の見積項目リスト

        Returns:
            単価と根拠情報が付与された見積項目リスト
        """
        logger.info(f"Enriching {len(items)} items with price RAG")

        enriched_items = []

        for item in items:
            # 同じ工事区分のKBエントリを検索
            matching_refs = [
                ref for ref in self.price_kb
                if ref.discipline == item.discipline
            ]

            # 項目名と仕様で簡易マッチング（実運用ではベクトル検索を使用）
            best_match = None
            best_score = 0.0

            for ref in matching_refs:
                # 簡易的な類似度計算（実運用ではembedding使用）
                score = 0.0
                if item.name and ref.description:
                    if item.name in ref.description or ref.description in item.name:
                        score += 0.5

                if item.specification and ref.features.get("specification"):
                    if item.specification == ref.features["specification"]:
                        score += 0.5

                if item.unit == ref.unit:
                    score += 0.3

                if score > best_score:
                    best_score = score
                    best_match = ref

            # 単価を設定
            if best_match and best_score >= 0.3:
                item.unit_price = best_match.unit_price
                item.price_references = [best_match.item_id]
                item.source_type = "rag"
                item.source_reference = f"KB:{best_match.item_id}(score={best_score:.2f})"

                # 金額を計算
                if item.quantity and item.unit_price:
                    item.amount = item.quantity * item.unit_price

                logger.debug(f"Matched '{item.name}' with '{best_match.description}' (score={best_score:.2f}, price=¥{best_match.unit_price:,})")
            else:
                logger.debug(f"No match found for '{item.name}' (best_score={best_score:.2f})")

            enriched_items.append(item)

        # 統計情報
        matched_count = sum(1 for item in enriched_items if item.unit_price is not None)
        logger.info(f"Matched {matched_count}/{len(enriched_items)} items with KB prices")

        return enriched_items


if __name__ == "__main__":
    # テスト実行
    kb_builder = PriceKBBuilder()

    # 見積書PDFから過去見積KBを構築
    estimate_pdf = "test-files/250918_送付状　見積書（都市ｶﾞｽ).pdf"

    if Path(estimate_pdf).exists():
        price_refs = kb_builder.extract_estimate_from_pdf(estimate_pdf)
        print(f"\n✅ 過去見積KB構築完了:")
        print(f"   抽出項目数: {len(price_refs)}")

        # KBを保存
        kb_output = "kb/price_kb.json"
        os.makedirs("kb", exist_ok=True)
        kb_builder.save_kb_to_json(price_refs, kb_output)
        print(f"   保存先: {kb_output}")

        # サンプル表示
        print(f"\n【価格KB サンプル（最初の5項目）】")
        for ref in price_refs[:5]:
            print(f"  - {ref.description} {ref.features.get('specification', '')} : ¥{ref.unit_price:,}/{ref.unit}")
    else:
        print(f"❌ 見積書が見つかりません: {estimate_pdf}")
